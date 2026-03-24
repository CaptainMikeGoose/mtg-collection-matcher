from pathlib import Path
from openpyxl import load_workbook
import re

def normalize(s: str) -> str:
    return " ".join(str(s).strip().split()).lower()

def base_name(name: str) -> str:
    # Removes " (anything)" at the end or anywhere in the name.
    # Example: "Swamp (Foil)" -> "Swamp"
    return re.sub(r"\s*\([^)]*\)", "", name).strip()

def parse_decklist_line(line: str):
    line = line.strip()
    if not line:
        return None

    parts = line.split(maxsplit=1)
    if len(parts) == 1:
        return parts[0].strip()

    if parts[0].isdigit():
        return parts[1].strip()

    return line

def read_decklist(txt_path: str) -> list[str]:
    lines = Path(txt_path).read_text(encoding="utf-8").splitlines()
    names = []
    for line in lines:
        name = parse_decklist_line(line)
        if name:
            names.append(name)
    return names

def main():
    #Settings
    deck_path = "decklist.txt"
    xlsx_path = "MtG Cards.xlsx"
    sheet_name = None
    name_col = "B"
    qty_col = "C"
    status_col = "D"   # must be empty to be available

    deck_cards = read_decklist(deck_path)

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Build lookup: normalized base name -> list of matching sheet rows
    # Each row stored as: (original_name_in_B, qty_in_C, status_in_D)
    sheet = {}

    for r in range(1, ws.max_row + 1):
        name_cell = ws[f"{name_col}{r}"].value
        if name_cell is None:
            continue

        name_str = str(name_cell).strip()
        if not name_str:
            continue

        # Skip likely set header rows like "KALADESH"
        if name_str.isupper() and len(name_str.split()) <= 5:
            continue

        status_val = ws[f"{status_col}{r}"].value
        status_str = "" if status_val is None else str(status_val).strip()

        # Only include AVAILABLE cards (column D empty)
        if status_str:
            continue

        qty_val = ws[f"{qty_col}{r}"].value
        qty = int(qty_val) if isinstance(qty_val, (int, float)) else None

        key = normalize(base_name(name_str))  # "Swamp (Foil)" and "Swamp" both map to "swamp"
        sheet.setdefault(key, []).append((name_str, qty, status_str))

    matched = []
    missing = []

    for card in deck_cards:
        key = normalize(base_name(card))
        if key in sheet:
            matched.append((card, sheet[key]))  # list of variants
        else:
            missing.append(card)

    print(f"Decklist cards: {len(deck_cards)}")
    print(f"Matched (available): {len(matched)}")
    print(f"Not found (or only unavailable): {len(missing)}\n")

    if matched:
        print("=== Matches (Available Variants) ===")
        for deck_name, options in matched:
            print(deck_name)
            for sheet_name, qty, status in options:
                print(f"  - {sheet_name}  |  qty: {qty if qty is not None else '?'}")

    if missing:
        print("\n=== Not Found / Not Available ===")
        for card in missing:
            print(card)

if __name__ == "__main__":
    main()